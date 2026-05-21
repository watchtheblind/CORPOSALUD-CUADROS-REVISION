from openpyxl import load_workbook
from core.configs_loader import clean_text


class WorkbookReader:
    """Lee el libro de carga y extrae encabezados + datos."""

    def __init__(self, path: str):
        self.path = path
        self.headers: tuple = ()
        self.rows: list = []
        self.idx: dict = {}

    def _find_header_row(self, sheet, max_row):
        """Método de apoyo para encontrar el índice de la fila ancla."""
        for row_idx, row_values in enumerate(sheet.iter_rows(max_row=max_row, values_only=True), 1):
            if any(clean_text(cell) == "CEDULA" for cell in row_values if cell):
                return row_idx
        return None

    def read_workbook(self):
        """
            Busca la fila con 'CEDULA' en las primeras 15 filas
            de cualquier hoja y carga todo desde ahí.
        """
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        raw_data_matrix = None
        SCAN_LIMIT = 15

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 1. Buscamos el número de fila que contiene el ancla
            header_row_number = self._find_header_row(sheet, SCAN_LIMIT)
            
            if header_row_number:
                # 2. Extraemos todo desde esa fila en adelante
                raw_data_matrix = list(sheet.iter_rows(min_row=header_row_number, values_only=True))
                break

        workbook.close()

        if not raw_data_matrix:
            raise Exception(
                f"No se encontró la columna 'CEDULA' en las primeras {SCAN_LIMIT} filas "
                f"de ninguna hoja. No se puede continuar."
            )
            
        self.headers = raw_data_matrix[0]
        self.rows = raw_data_matrix[1:]
        
        self.column_index_map = {
            clean_text(header_name): index
            for index, header_name in enumerate(self.headers) 
            if header_name
        }

        return self