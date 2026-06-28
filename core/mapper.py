from core.configs_loader import clean_text, load_mapping, load_concepts_with_factors
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill

VERDE = PatternFill(
    start_color="92D050",
    end_color="92D050",
    fill_type="solid"
)


class Mapper:
    """
    Resuelve qué columna de la carga va a qué columna de la plantilla.
    """

    def __init__(self, idx_wb: dict, ws_template: Worksheet):
        self.idx_wb = idx_wb
        self.ws_p = ws_template
        self.mapping = load_mapping()
        self.factor_concepts = [
            clean_text(concept) for concept in load_concepts_with_factors()
        ]

        # Índices de la plantilla
        self.idx_template = {
            clean_text(ws_template.cell(1, c).value): c
            for c in range(1, ws_template.max_column + 1)
            if ws_template.cell(1, c).value
        }

    def build_template_mapping(self) -> list[tuple[int, int]]:
        """
        Coordina la vinculación entre columnas del archivo de carga y la plantilla.
        Retorna una lista de tuplas: (indice_plantilla, indice_carga).
        """
        column_mapping = []

        for identifier, (workbook_column_name, template_column_name) in self.mapping.items():
            # Buscamos los índices en los diccionarios de índices previamente cargados
            index_load = self.idx_wb.get(clean_text(workbook_column_name))
            template_index = self.idx_template.get(clean_text(template_column_name))

            # Cláusula de guarda: Si uno de los dos no existe, no hay puente posible
            if index_load is None or template_index is None:
                continue

            # Registramos el vínculo
            column_mapping.append((template_index, index_load))

            # Tarea estética: Marcamos visualmente las columnas encontradas
            self._highlight_mapped_headers(template_index, template_column_name)

        return column_mapping

    def _highlight_mapped_headers(self, columna_idx: int, nombre_plantilla: str):
        """Aplica formato visual a la plantilla para las columnas procesadas."""
        # Pintamos la columna principal
        self.ws_p.cell(row=1, column=columna_idx).fill = VERDE

        # Si el concepto requiere un factor (columna adyacente), también la pintamos
        es_concepto_con_factor = clean_text(nombre_plantilla) in self.factor_concepts
        if es_concepto_con_factor:
            self.ws_p.cell(row=1, column=columna_idx + 1).fill = VERDE

    def _find_best_formula_row(self, col_indices: list, start_row: int, depth: int) -> int | None:
        """
        Escanea hacia abajo un rango limitado de filas para encontrar dónde 
        están las fórmulas realmente. Retorna el índice de la primera fila válida.
        """
        for current_row in range(start_row, start_row + depth):
            for column_index in col_indices:
                value = self.ws_p.cell(row=current_row, column=column_index).value
                if value and str(value).startswith("="):
                    return current_row
        return None


    def get_formulas(self) -> list[tuple[int, str, str]]:
        """
        Identifica columnas de cálculo que contienen una fórmula de Excel.
        Busca dinámicamente la fila donde comienzan los datos si no están en la fila inicial.
        """
        detected_formulas = []
        TARGET_HEADERS = {"FORMULA", "DIFERENCIA"}
        
        # Configuraciones de búsqueda fáciles de modificar
        HEADER_ROW = 1
        START_SEARCH_ROW = 2  # Fila donde esperamos que empiecen las fórmulas
        MAX_SEARCH_DEPTH = 15 # Cuántas filas hacia abajo buscar si la START_SEARCH_ROW está vacía
        
        # 1. Identificamos qué columnas coinciden con nuestros encabezados objetivo
        # Guardamos los índices para no procesar columnas irrelevantes
        #WS_p = WORKSHEET PROCESS o hoja activa
        target_column_indices = [
            column_index for column_index in range(1, self.ws_p.max_column + 1)
            if (header := self.ws_p.cell(row=HEADER_ROW, column=column_index).value)
            and clean_text(header) in TARGET_HEADERS
        ]

        if not target_column_indices:
            return []

        # 2. Buscamos la "Fila Modelo" (la primera que tenga una fórmula en esas columnas)
        # Esto ahorra procesador al no repetir la búsqueda por cada columna.
        formula_row_index = self._find_best_formula_row(target_column_indices, START_SEARCH_ROW, MAX_SEARCH_DEPTH)
        
        if not formula_row_index:
            return []

        # 3. Extraemos los datos finales usando la fila encontrada
        for column_index in target_column_indices:
            model_cell = self.ws_p.cell(row=formula_row_index, column=column_index)
            formula_content = model_cell.value
            
            if formula_content and str(formula_content).startswith("="):
                detected_formulas.append((
                    column_index, 
                    formula_content, 
                    model_cell.coordinate
                ))
                
        return detected_formulas

    def has_adjacent_factor(self, col_plantilla: int) -> bool:
        """Verifica si la columna tiene un factor a su derecha."""
        header = clean_text(self.ws_p.cell(1, col_plantilla).value)
        return header in self.factor_concepts