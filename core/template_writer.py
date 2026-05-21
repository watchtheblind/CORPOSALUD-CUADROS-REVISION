from datetime import datetime, date
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formula.translate import Translator
from core.configs_loader import clean_text
from core.mapper import Mapper

def parse_latin_number(value: any) -> float | int | any:
    """
    Normaliza y convierte valores a tipos numéricos, manejando formatos latinos.
    
    Lógica de normalización:
    1. Si tiene '.' (miles) y ',' (decimales) -> Elimina '.' y cambia ',' por '.'
    2. Si solo tiene ',' -> Cambia ',' por '.'
    3. En cualquier otro caso -> Intenta conversión directa.
    """
    # Escenario 1: El valor ya es numérico o está vacío
    if value in (None, ""):
        return None
        
    if isinstance(value, (int, float)):
        return value

    # Escenario 2: Limpieza de strings
    text_value = str(value).strip()
    
    # Aplicamos transformación según separadores presentes
    normalized_text = _normalize_separators(text_value)

    try:
        # Intentamos la conversión final
        return float(normalized_text)
    except (ValueError, TypeError):
        # Si falla (ej. es un texto como "ABC"), devolvemos el valor original
        return value

def _normalize_separators(text: str) -> str:
    """
    Detecta y ajusta los separadores de miles y decimales.
    """
    has_comma = "," in text
    has_dot = "." in text

    # Formato: 1.234,56 (Miles con punto, decimal con coma)
    if has_comma and has_dot:
        return text.replace(".", "").replace(",", ".")
    
    # Formato: 1234,56 (Solo coma decimal)
    if has_comma:
        return text.replace(",", ".")
        
    return text

# Columnas que NO se convierten a número
TEXT_COLUMNS = {"CEDULA", "CUENTANOMINA"}

# parts que forman "APELLIDOS Y NOMBRES"
NAME_PARTS = ["1ER APELLIDO", "2DO APELLIDO", "1ER NOMBRE", "2DO NOMBRE"]


class TemplateWriter:
    """Escribe los datos de la carga en la plantilla."""

    def __init__(
        self,
        ws: Worksheet,
        mapper: Mapper,
        load_index: dict,
    ):
        self.ws = ws
        self.mapper = mapper
        self.load_index = load_index

        # Pre-calcular índices útiles
        self.plan = mapper.build_template_mapping()
        self.formulas = mapper.get_formulas()
        self.colum_names = mapper.idx_plantilla.get(
            clean_text("APELLIDOS Y NOMBRES")
        )
        self.idx_NAME_PARTS = [
            load_index.get(clean_text(n)) for n in NAME_PARTS
        ]
        self.idx_cedula = load_index.get(clean_text("CEDULA"))

    def clean_previous_data(self):
        """Borra celdas que no sean fórmula desde la fila 3."""
        if self.ws.max_row < 3:
            return
        for row in self.ws.iter_rows(min_row=3, max_row=self.ws.max_row):
            for cell in row:
                if cell.value and not str(cell.value).startswith('='):
                    cell.value = None

    def write_everything(self, filas: list):
        """Escribe todas las filas de datos en la plantilla."""
        self.clean_previous_data()

        for r_off, fila in enumerate(filas, 3):
            if not self.valid_row(fila):
                continue

            self._write_data(r_off, fila)
            self._write_full_name(r_off, fila)
            self._write_formulas(r_off)

    # --- Métodos privados ---

    def valid_row(self, fila) -> bool:
        """Verifica que la fila tenga una cédula válida."""
        if self.idx_cedula is None:
            return False
        cedula = str(fila[self.idx_cedula]).strip() if fila[self.idx_cedula] else ""
        return cedula.lower() not in ["", "none", "total", "totales"]

    def _write_data(self, r_off: int, fila: tuple):
        """Escribe cada celda según el plan de trabajo."""
        for cp, cc in self.plan:
            val = fila[cc]

            # Conversión de tipo
            header = clean_text(self.ws.cell(1, cp).value)
            if header not in TEXT_COLUMNS:
                val = parse_latin_number(val)

            # Fechas
            if isinstance(val, datetime):
                val = val.date()
            if isinstance(val, (date, datetime)):
                self.ws.cell(r_off, cp).number_format = 'DD/MM/YYYY'

            self.ws.cell(r_off, cp).value = val

            # Factor adyacente
            if self.mapper.tiene_factor_adyacente(cp):
                self._copiar_factor(r_off, cp, cc, fila)

    def _copiar_factor(self, r_off: int, cp: int, cc: int, fila: tuple):
        """Copia el value de la columna adyacente derecha (factor)."""
        try:
            value_factor = fila[cc + 1]
            if value_factor is not None:
                self.ws.cell(
                    row=r_off,
                    column=cp + 1,
                    value=parse_latin_number(value_factor)
                )
        except IndexError:
            pass

    def _write_full_name(self, r_off: int, fila: tuple):
        """Concatena las partes del nombre en la columna correspondiente."""
        if not self.colum_names:
            return
        parts = [
            str(fila[i]).strip()
            for i in self.idx_NAME_PARTS
            if i is not None and fila[i]
        ]
        self.ws.cell(r_off, self.colum_names).value = " ".join(parts).upper()

    def _write_formulas(self, r_off: int):
        """Traduce y escribe fórmulas para la fila actual."""
        for col, model_formula, coord_origen in self.formulas:
            destination_coord = self.ws.cell(r_off, col).coordinate
            self.ws.cell(r_off, col).value = Translator(
                model_formula, origin=coord_origen
            ).translate_formula(destination_coord)