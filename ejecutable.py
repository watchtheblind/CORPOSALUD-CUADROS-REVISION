import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formula.translate import Translator
import os
import re
import sys
import threading
from datetime import datetime, date
import threading
# Importamos la herramienta de carga desde tu utils.py
from utils import ejecutar_tarea_con_carga, CargaUI

# --- 1. CONFIGURACIÓN DE MAPEO ---
MAPEO_COLUMNAS = {
    "CEDULA": ["CEDULA", "CEDULA"],
    "SEXO": ["SEXO", "SEXO"],
    "DESC_CARGO": ["DESC. CARGO", "DESC. CARGO"],
    "UBIC_ADMIN": ["UBICACION ADMINISTRATIVA", "UBICACION ADMINISTRATIVA (ASIC)"],
    "COD_RAC": ["COD. CARGO", "CODIGO RAC"],
    "NIVEL": ["NIVEL CARGO", "NIVEL CARGO"],
    "GRADO": ["GRADO", "GRADO"],
    "HORAS": ["CARGA HORARIA", "CARGA HORARIA"],
    "ESP_MED": ["ESPECIALIDAD MÉD 1", "ESPECIALIDAD DEL MEDICO"],
    "UBIC_FISICA": ["UBICACION FISICA", "UBICACIÓN FISICA"],
    "CENTRO": ["NOMBRE CENTRO", "NOMBRE CENTRO"],
    "ANNOS_REC": ["ANNOS REC", "ANNOS REC"],
    "MESES_REC": ["MESES REC", "MESES REC"],
    "F_INGRESO": ["FECHA INGRESO", "FECHA INGRESO"],
    "F_NAC": ["FECHA NAC.", "FECHA NAC."],
    "BANCO": ["ENTIDAD BANCARIA", "ENTIDAD BANCARIA"],
    "CUENTA": ["CUENTA NOMINA", "CUENTA NOMINA"],
    "CTA_ACTIVA": ["CUENTA ACTIVA", "CUENTA ACTIVA"],
    "TIPO_EMP": ["TIPO EMPLEADO", "TIPO EMPLEADO"],
    "F_INACTIVA": ["FECHA INACTIVACIÓN", "FECHA DE CTA INACTIVA"],
    "P_ANTIG": ["PORC ANTIGUEDAD", "PORC ANTIGUEDAD"],
    "P_ESCAL": ["PORC ESCALAF", "PORC ESCALAF"],
    "HIJOS": ["NRO HIJOS", "NUMERO DE HIJOS"],
    "SUELDO_NORM": ["SUELDO NORMAL", "SUELDO NORMAL"],
    "C101": ["101 SUELDO Y/O SALARIO", "101 SUELDO Y/O SALARIO"],
    "C117": ["117 DIF. DE SUELDO Y/O SALARIO", "117 DIF. DE SUELDO Y/O SALARIO"],
    "C102": ["102 COMPENSACION POR EVALUACIÓN", "102 COMPENSACION POR EVALUACIÓN"],
    "C111": ["111 COMP. Y AJUSTE DE ESCALA", "111 COMP. Y AJUSTE DE ESCALA"],
    "C106": ["106 PRIMA ANTIGUEDAD", "106 PRIMA ANTIGUEDAD"],
    "C110": ["110 AJUST. ESCALAFÓN MÉDICO", "110 AJUST. ESCALAFÓN MÉDICO"],
    "C110_8": ["110.8 ESCALAFÓN", "110.8 ESCALAFÓN"],
    "C400": ["400 PRIMA SIST. PUB. SALUD", "400 PRIMA SIST. PUB. SALUD"],
    "C105_4": ["105.4 PRIMA PROFE. 20%", "105.4 PRIMA PROFE. 20%"],
    "C105_5": ["105.5 PRIMA PROFE. 25%", "105.5 PRIMA PROFE. 25%"],
    "C105_6": ["105.6 PRIMA PROFE. 30%", "105.6 PRIMA PROFE. 30%"],
    "C105_7": ["105.7 PRIMA PROFE. 35%", "105.7 PRIMA PROFE. 35%"],
    "C105_8": ["105.8 PRIMA PROFE. 40%", "105.8 PRIMA PROFE. 40%"],
    "C111_5": ["111.5 COMP. EVAL. 2DO SEM 2023", "111.5 COMP. EVAL. 2DO SEM 2023"],
    "C124_4": ["124.4 DIA ADICIONAL", "124.4 DIA ADICIONAL"],
    "C140_2_16": ["140.2.16 DOMINGO Y FERIADO NOCT", "140.2.16 DOMINGO Y FERIADO NOCT"],
    "F140_2_16": ["FACTOR DOM Y FER NOCT", "FACTOR DOM Y FER NOCT"],
    "C140_2_17": ["140.2.17 DOMINGO Y FERIADO 24H", "140.2.17 DOMINGO Y FERIADO 24H"],
    "F140_2_17": ["FACTOR DOM Y FER 24H", "FACTOR DOM Y FER 24H"],
    "C140_2_18": ["140.2.18 DOMINGO Y FER. DIURNOS", "140.2.18 DOMINGO Y FER. DIURNOS"],
    "F140_2_18": ["FACTOR DOM Y FER. DIURNOS", "FACTOR DOM Y FER. DIURNOS"],
    "C140_1": ["140.1 DOM Y FER MED", "140.1 DOMINGO Y FERIADO MEDICO"],
    #"F140_1": ["FACTOR DOM Y FER MED", "FACTOR"],
    "C120_5_1": ["120.5.1 BONO NOCT. FIJO", "120.5.1 BONO NOCT. FIJO"],
    "C143_4_10": ["143.4.10 BONO NOCT. TEMP", "143.4.10 BONO NOCT. TEMP (6 HRS)"],
    "C143_5": ["143.5 BONO NOCT ADMIN", "143.5 BONO NOCT. 8 H"],
    "C143_3": ["143.3 BONO NOCTURNO MED (6HORAS)", "143.3 BONO NOCTURNO MEDICO (6HORAS)"],
    "C143_4": ["143.4 BONO NOCTURNO MED (8HORAS)", "143.4 BONO NOCTURNO MEDICO (8HORAS)"],
    "C160_17": ["160.17 BONO VACACIONAL", "160.17 BONO VACACIONAL"],
    "C128": ["128 BECAS", "128 BECAS"],
    "C155_3": ["155.3 PRIMA POR HIJOS", "155.3 PRIMA POR HIJOS"],
    "C155_5": ["155.5 DIA DEL PADRE", "155.5 DIA DEL PADRE"],
    "C155_6": ["155.6 DIA DEL NIÑO", "155.6 DIA DEL NIÑO"],
    "C148_1": ["148.1 UNIFORMES", "148.1 UNIFORMES"],
    "C161": ["161 BONO NAVIDEÑO", "161 BONO NAVIDEÑO"],
    "C127_1": ["127.1 JUGUETES", "127.1 JUGUETES"],
    "C927": ["927 AYUD SERV FUNER", "927 AYUD SERV FUNERARIO"],
    "C135": ["135 BONIF POR MATRIM", "135 BONIF POR MATRIMONIO"],
    "C132": ["132 BONIF POR NAC", "132 BONIF POR NACIMIENTO"],
    "C178": ["178 DESC PARC X PAGO INDEBID", "178 DESC PARC X PAGO INDEBIDO"],
    "C179": ["179 DESC. DIA(S) NO LABORADO", "179 DESC. DIA(S) NO LABORADO(S)"],
    "C202": ["202 JUZG PRIMER DE MENORES", "202 JUZG PRIMERO DE MENORES"],
    "C210": ["210 S.S.O. (4%)", "210 S.S.O. (4%)"],
    "C212": ["212 CAHORMINSAS", "212 CAHORMINSAS"],
    "C213": ["213 SUNEP", "213 SUNEP"],
    "C215": ["215 TRIBUNAL (PERMANENTE)", "215 TRIBUNAL (PERMANENTE)"],
    "C223": ["223 PREST CAJA DE AHORRO", "223 PREST CAJA DE AHORRO"],
    "C224": ["224 COLEGIO ENFERMERAS", "224 COLEGIO ENFERMERAS"],
    "C233": ["233 SINDICAT UNICO TRAB.DE LA SALUD", "233 SINDICAT UNICO TRAB.DE LA SALUD"],
    "C238": ["238 CAEMINSA", "238 CAEMINSA"],
    "C244": ["244 PERDIDA INVOLUNTARIA DEL EMPLEO", "244 PERDIDA INVOLUNTARIA DEL EMPLEO"],
    "C245": ["245 FAOV", "245 FAOV"],
    "C255": ["255 COLEGIO BIOANALISTA", "255 COLEGIO BIOANALISTA"],
    "C257": ["257 SOCIEDAD ANESTECIOLOGOS", "257 SOCIEDAD ANESTECIOLOGOS"],
    "C286": ["286 COLEGIO NUTRIC. Y DIET", "286 COLEGIO NUTRIC. Y DIET"],
    "C301": ["301 FONDO PENSIONES (JUBILACION)", "301 FONDO PENSIONES (JUBILACION)"],
    "C321": ["321 SINBOPROENF", "321 SINBOPROENF"],
    "C370": ["370 DEL REG INPRENFER ARAGUA", "370 DEL REG INPRENFER ARAGUA"],
    "C583": ["583 SISTRASALUD", "583 SISTRASALUD"],
    "C799": ["799 ASUNAJUPENSAPROSO", "799 ASUNAJUPENSAPROSO"],
    "C807": ["807 TRIB DE PROTEC DE NIÑOS", "807 TRIBUNAL DE PROTECCION DE NIÑOS"],
    "C978": ["978 SINDICATO OSBESS", "978 SINDICATO OSBESS"],
    "C513": ["513 FENISISTRASALUD", "513 FENISISTRASALUD"],
    "C581": ["581 SITRASSS MIRANDA", "581 SITRASSS MIRANDA"],
    "C603": ["603 COLEGIO ENFER. EDO MIRANDA", "603 COLEG ENF EDO MIRANDA"],
    "C964": ["964 SAPTRASEZ", "964 SAPTRASEZ"],
    "C965": ["965 COLEG ENF EDO LARA", "965 COLEG ENF EDO LARA"],
    "C966": ["966 SUPTRASADESINCA", "966 SUPTRASADESINCA"],
    "C967": ["967 COLEG ENF EDO CARABOBO", "967 COLEG ENF EDO CARABOBO"],
    "C973": ["973 SESTRASALUD", "973 SESTRASALUD"],
    "C907": ["907 CAJA DE AHORRO", "907 CAJA DE AHORRO"],
    "C910": ["910 PENSIÓN ALIMENTACIÓN", "910 PENSIÓN ALIMENTACIÓN"]
}

# --- 2. UTILIDADES ---
def convertir_num_fiel(valor):
    if valor is None or valor == "": return None
    if isinstance(valor, (int, float)): return valor
    try:
        s = str(valor).strip()
        if "," in s and "." not in s: s = s.replace(",", ".")
        elif "," in s and "." in s: s = s.replace(".", "").replace(",", ".")
        return float(s)
    except:
        return valor

# --- 3. CLASE PRINCIPAL ---
class ProcesadorNomina:
    def __init__(self):
        self.verde_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    def limpiar(self, t):
        return re.sub(r'[^A-Z0-9]', '', str(t).upper()) if t else ""

    def obtener_ruta(self, archivo):
        if hasattr(sys, '_MEIPASS'): return os.path.join(sys._MEIPASS, archivo)
        return os.path.join(os.path.abspath("."), archivo)

    def ejecutar(self):
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        ruta_plantilla = self.obtener_ruta("plantilla2.xlsx")
        if not os.path.exists(ruta_plantilla):
            messagebox.showerror("Error", f"No se encontró la plantilla en:\n{ruta_plantilla}")
            return

        ruta_carga = filedialog.askopenfilename(title="Seleccionar LIBRO CARGA", parent=root)
        if not ruta_carga:
            root.destroy()
            return

        # Modificamos para crear el objeto de carga manualmente y tener control
        loading = CargaUI(root, "Procesando Nómina y Fórmulas...")

        # Lanzamos la lógica en un hilo pasándole el objeto 'loading'
        threading.Thread(
            target=self.logica_procesamiento, 
            args=(ruta_carga, ruta_plantilla, root, loading),
            daemon=True
        ).start()
        
        root.mainloop()
    def procesar_factores_adyacentes(self, ws_p, r_off, cp, cc, fila_v):
            """
            Si la columna destino es un concepto que requiere factor, 
            copia el valor de la columna adyacente derecha del Libro Carga
            a la columna adyacente derecha de la Plantilla.
            """
            # Lista de encabezados que sabemos que tienen un factor a su derecha
            conceptos_con_factor = [
                "140.1 DOMINGO Y FERIADO MEDICO",
                "143.4.10 BONO NOCT. TEMP (6 HRS)",
                "143.5 BONO NOCT. 8 H",
                "143.3 BONO NOCTURNO MEDICO (6HORAS)",
                "143.4 BONO NOCTURNO MEDICO (8HORAS)"
            ]
            
            # Obtenemos el nombre del encabezado en la plantilla para validar
            header_p = self.limpiar(ws_p.cell(1, cp).value)
            
            # Verificamos si este encabezado (limpio) coincide con nuestros conceptos
            for concepto in conceptos_con_factor:
                if self.limpiar(concepto) == header_p:
                    # El factor en Carga está en cc + 1
                    # El factor en Plantilla está en cp + 1
                    try:
                        valor_factor = fila_v[cc + 1]
                        if valor_factor is not None:
                            ws_p.cell(row=r_off, column=cp + 1, value=convertir_num_fiel(valor_factor))
                    except IndexError:
                        pass # Evita errores si no hay columna a la derecha
                    break
    def logica_procesamiento(self, ruta_carga, ruta_plantilla, root, loading):
        try:
            # 1. CARGA DE DATOS
            wb_c = load_workbook(ruta_carga, read_only=True, data_only=True)
            matriz = None
            for sn in wb_c.sheetnames:
                ws = wb_c[sn]
                for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
                    if any(self.limpiar(c) == "CEDULA" for c in row if c):
                        matriz = list(ws.iter_rows(min_row=r_idx, values_only=True))
                        break
                if matriz: break
            wb_c.close()

            if not matriz: raise Exception("No se encontró la columna 'CEDULA' en el archivo de carga.")

            # 2. MAPEO DE ÍNDICES
            header_c = matriz[0]
            idx_c = {self.limpiar(v): i for i, v in enumerate(header_c) if v}
            
            wb_p = load_workbook(ruta_plantilla)
            ws_p = wb_p.active
            idx_p = {self.limpiar(ws_p.cell(1, c).value): c for c in range(1, ws_p.max_column + 1) if ws_p.cell(1, c).value}

            # 3. PLAN DE TRABAJO
            plan_trabajo = []
            for clave, nombres in MAPEO_COLUMNAS.items():
                c_orig = idx_c.get(self.limpiar(nombres[0]))
                c_dest = idx_p.get(self.limpiar(nombres[1]))
                if c_orig is not None and c_dest is not None:
                    plan_trabajo.append((c_dest, c_orig))
                    ws_p.cell(1, c_dest).fill = self.verde_fill
                    header_limpio = self.limpiar(nombres[1])
                    conceptos_con_factor = [self.limpiar(c) for c in [
                        "140.1 DOMINGO Y FERIADO MEDICO",
                        "143.4.10 BONO NOCT. TEMP (6 HRS)",
                        "143.5 BONO NOCT. 8 H",
                        "143.3 BONO NOCTURNO MEDICO (6HORAS)",
                        "143.4 BONO NOCTURNO MEDICO (8HORAS)"
                    ]]
                    if header_limpio in conceptos_con_factor:
                        ws_p.cell(1, c_dest + 1).fill = self.verde_fill

            # 4. COLUMNAS ESPECIALES
            col_nombres = idx_p.get(self.limpiar("APELLIDOS Y NOMBRES"))
            idx_nom_partes = [idx_c.get(self.limpiar(n)) for n in ["1ER APELLIDO", "2DO APELLIDO", "1ER NOMBRE", "2DO NOMBRE"]]
            
            formulas = []
            for c in range(1, ws_p.max_column + 1):
                h_val = ws_p.cell(1, c).value
                if h_val and self.limpiar(h_val) in ["FORMULA", "DIFERENCIA"]:
                    f_m = ws_p.cell(2, c).value
                    if f_m and str(f_m).startswith("="):
                        formulas.append((c, f_m, ws_p.cell(2, c).coordinate))

            # 5. ESCRITURA MASIVA
            # --- Escudo 1: Limpieza de datos viejos (respetando fórmulas) ---
            if ws_p.max_row >= 3:
                for row in ws_p.iter_rows(min_row=3, max_row=ws_p.max_row):
                    for cell in row:
                        # Solo borramos si NO es una fórmula (las fórmulas empiezan por '=')
                        if cell.value and not str(cell.value).startswith('='):
                            cell.value = None

            # --- Escudo 2: Filtro de filas basura en la carga ---
            idx_cedula_carga = idx_c.get(self.limpiar("CEDULA"))
            for r_off, fila_v in enumerate(matriz[1:], 3):
                # Verificamos si la fila actual tiene una cédula válida
                cedula_check = str(fila_v[idx_cedula_carga]).strip() if idx_cedula_carga is not None else ""
                
                # Si la cédula está vacía, es None o dice "TOTAL", paramos el proceso de esa fila
                if not cedula_check or cedula_check.lower() in ["none", "", "total", "totales"]:
                    continue
                for cp, cc in plan_trabajo:
                    val = fila_v[cc]
                    header_actual = self.limpiar(ws_p.cell(1, cp).value)
                    if header_actual not in ["CEDULA", "CUENTANOMINA"]:
                        val = convertir_num_fiel(val)
                    if isinstance(val, datetime): val = val.date()
                    if isinstance(val, (date, datetime)):
                        ws_p.cell(r_off, cp).number_format = 'DD/MM/YYYY'
                    ws_p.cell(r_off, cp).value = val
                    self.procesar_factores_adyacentes(ws_p, r_off, cp, cc, fila_v)

                if col_nombres:
                    partes = [str(fila_v[i]).strip() for i in idx_nom_partes if i is not None and fila_v[i]]
                    ws_p.cell(r_off, col_nombres).value = " ".join(partes).upper()

                for cf, fm, orig in formulas:
                    dest = ws_p.cell(r_off, cf).coordinate
                    ws_p.cell(r_off, cf).value = Translator(fm, origin=orig).translate_formula(dest)

            # 6. FINALIZACIÓN
            root.after(0, lambda: self.finalizar_proceso(wb_p, root, loading))

        except Exception as e:
            root.after(0, lambda: messagebox.showerror("Error", str(e)))
            root.after(0, loading.cerrar)

    def finalizar_proceso(self, wb_p, root, loading):
        ruta_s = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"Nomina_Procesada_{date.today().strftime('%d_%m_%Y')}",
            title="Guardar Resultado Final",
            parent=root
        )
        
        if ruta_s:
            # Actualizamos mensaje de la carga
            for widget in loading.top.winfo_children():
                if isinstance(widget, tk.Label):
                    widget.config(text="Escribiendo archivo en disco...\nPor favor, no cierre el programa.")

            def guardar_fisicamente():
                try:
                    wb_p.save(ruta_s)
                    root.after(0, lambda: self.cerrar_con_exito(loading, ruta_s, root))
                except Exception as e:
                    root.after(0, lambda: messagebox.showerror("Error al guardar", str(e)))
                    root.after(0, loading.cerrar)

            threading.Thread(target=guardar_fisicamente, daemon=True).start()
        else:
            loading.cerrar()
            root.destroy()

    def cerrar_con_exito(self, loading, ruta, root):
        loading.cerrar()
        messagebox.showinfo("Éxito", "El archivo se ha guardado correctamente.")
        os.startfile(ruta)
        root.destroy()

if __name__ == "__main__":
    ProcesadorNomina().ejecutar()